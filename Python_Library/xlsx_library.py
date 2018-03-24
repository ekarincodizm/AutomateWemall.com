import xlrd
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

def set_sku_and_qty_in_file_excel(file_path,content_list):
    wb = load_workbook(file_path)
    # get all sheet name
    all_sheet = wb.get_sheet_names()
    # delete all sheet
    for sheet in all_sheet:
        ws=wb.get_sheet_by_name(sheet)
        wb.remove_sheet(ws)
    # create the new one
    ws = wb.create_sheet(title='v_stock', index=0)
    ws = wb.active
    # set data content to excel file
    # content_list = [['A1', 'SKU'], ['B1', 'QTY'], ['A2', 'A1111AAWW0001'], ['B2', 100]]
    for content in content_list:
        try:
            value = float(content[1])
        except ValueError:
            value = (content[1])
        ws[content[0]] = value
    wb.save(file_path)

def verify_example_template_of_mass_upload_virtual_stock(excel_filename):

    COLUMN_SKU     = 0
    COLUMN_QTY     = 1

    expected_data = [['SKU', 'QTY']]

    try:
        if not excel_filename:
            excel_filename = "../Resource/TestData/merchant/upload_stock_excel_template.xlsx"

        book = xlrd.open_workbook(excel_filename)
        first_sheet = book.sheet_by_index(0)
        nrows=first_sheet.nrows
        ncols=first_sheet.ncols

        if nrows < 1:
            raise ValueError('The example excel file contains no data.')

        count = 0
        for row in range(0,nrows):
            SKU=first_sheet.cell(row, COLUMN_SKU).value
            QTY=first_sheet.cell(row, COLUMN_QTY).value

            if SKU != expected_data[count][0]:
                raise ValueError('SKU not match - expect ' + expected_data[count][0] + ' actual ' + SKU)

            if QTY != expected_data[count][1]:
                raise ValueError('QTY not match - expect ' + expected_data[count][1] + ' actual ' + QTY)

            count = count + 1
        return "True"
    except Exception as e:
        print e
        print "Verify example excel failed."
        return "False"

def set_sku_and_price_in_file_excel(file_path,content_list):
    wb = load_workbook(file_path)
    # get all sheet name
    all_sheet = wb.get_sheet_names()
    # delete all sheet
    for sheet in all_sheet:
        ws=wb.get_sheet_by_name(sheet)
        wb.remove_sheet(ws)
    # create the new one
    ws = wb.create_sheet(title='v_price', index=0)
    ws = wb.active
    # set data content to excel file
    # content_list = [['A1', 'SKU'], ['B1', 'Normal price'], ['C1', 'Special price'], ['A2', 'A1111AAWW0001'], ['B2', 500], ['C3', 400]]
    for content in content_list:
        try:
            value = float(content[1])
        except ValueError:
            value = (content[1])
        ws[content[0]] = value
        print content

    wb.save(file_path)


def read_price_in_history(file_path):
    data = []
    wb = load_workbook(file_path)
    mySheet = wb['sheet1']
    key = 0
    for sheet in mySheet:
        data.append(sheet[0].value.replace('.0', ''))
        data.append(sheet[1].value.replace('.0', ''))
        data.append(sheet[2].value.replace('.0', ''))
        data.append(sheet[3].value)
    print data
    return data
