import xlrd
import xlsxwriter
import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from helper import *

# function to open an Excel File
def open_excel_file(path):    
    # Open and read an Excel file
    book = xlrd.open_workbook(path)
 
    first_sheet = book.sheet_by_index(0)
    nrows=book.sheet_by_index(0).nrows
    ncols=book.sheet_by_index(0).ncols

    lst=[]
    for row in range(0,nrows):
        for col in range(0,ncols):
            cell=book.sheet_by_index(0).cell(row,col)
            lst.append(cell.value)
    return lst

# function to compare two lists
def compare_data(list1, list2):
    if(len(list1)==len(list1)):
        for i in range(0,len(list1)):
            if list1[i] != list2[i]:
                return False
        return True
    else:
        return False

# function to compare content of two excel books
def compare_excel_content_data(excel1_file,excel2_file):
    content1 = open_excel_file(excel1_file)
    content2 = open_excel_file(excel2_file)
    result = compare_data(content1,content2)
    assert result, 'Contents are different'
	
def py_write_data_to_excel(data=None) : 
    wb = openpyxl.Workbook()
    sheet = wb.active

    BuiltIn().log_to_console(data)

    i = 1
    sheet['A1'] = "Product Pkey"
    for product in data : 
        #BuiltIn().log_to_console("product_pkey = " + str(product["product_pkey"]))
        
        
        sheet['A' + str(i + 1)] = product['product_pkey']
        print "bbb"
        
        i = i + 1
        BuiltIn().log_to_console(i)

    # print sheet

    path = get_canonical_path("../../../Resource/TestData/Alias/template_mass_alias_upload.csv")
    
    wb.save(path)



# 2106374239768
# 2804434080543
# 2782548680360
# 2545332992126

def group(lst, n):
    return zip(*[lst[i::n] for i in range(n)])

def create_and_write_to_excel_file(filename, content_list):
    # Create an new Excel file and add a worksheet.

    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()

    #content_list=[1,1,'hello',2,1,'brother',3,1,'how are you',4,1,'are you good today']
    t = group(content_list, 3)
    for item in t:
        worksheet.write(int(item[0]), int(item[1]), item[2])

    # close work book
    workbook.close()

def create_excel_file_vip_email(filename, email):
    # Create an new Excel file and add a worksheet.
    # xfile = openpyxl.load_workbook(filename)
    # sheet = xfile.get_sheet_by_name('Sheet1')
    # sheet['A2'] = email
    # xfile.save(filename)
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    # worksheet.write_formula('A1', 'VIP Email')
    worksheet.write_formula('A2', email)
    workbook.set_calc_mode('manual')
    # close work book
    workbook.close()

def write_excel_for_upload_operation_status(file_name, content_list) : 
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()

